#!/usr/bin/python
#-*- coding: utf-8 -*-

import sys
import time
import urllib
import urllib.parse
import requests
import numpy as np
import imp
from bs4 import BeautifulSoup
from openpyxl import Workbook
from urllib import request
from urllib import error
from fake_useragent import UserAgent
imp.reload(sys)

# Fake User Agents
ua = UserAgent()
people=[['No.','Name','Company']]

def builder_spider():
    page_num=1;
    builder_list=[['No.','Name','Reg No','Areas','Contact','T','F','Address','Web','Key Projects','Awards']]
    
    while(page_num>0):
        url='https://www.mbawa.com/become-a-member/find-a-member/page/'+str(page_num)+'/?member_type=residential-builder&button&find-a-member=true'
        time.sleep(np.random.rand()*5) # Random waitling time to cover
        
        # Try to get html code
        try:
            req = request.Request(url, headers={ua.random})
            source_code = request.urlopen(req).read()
            plain_text=str(source_code).replace('<br>',' ').replace('<br/>',' ').replace('<br />',' ')
        except (error.HTTPError, error.URLError) as e: 
            print_builder_lists_excel(builder_list,page_num)
            print ('Access denied.')
            continue
        
        soup = BeautifulSoup(plain_text)
        list_soup = soup.find('ol', {'class': 'results table-list'})
        
        if list_soup == None:
            page_num=-1
            print_builder_lists_excel(builder_list,page_num)
            print('All result has been stored.')
            break

        for builder_info in list_soup.findAll('article'):
            name = builder_info.find('a', {'target':'_blank'}).string.strip()
            try:
                company_url = builder_info.find('a', {'target':'_blank'}).get('href')
                detail = get_detail(company_url, name)
            except:
                print('No detail get')
                detail=[''] * 9
            builder_list.append([len(builder_list),name,detail[0],detail[1],detail[2],detail[3],detail[4],detail[5],detail[6],detail[7],detail[8]])
        print ('Downloading Information From Page %d' % page_num)
        page_num+=1
    return 1


def get_detail(url, name):
    detail=[''] * 9
    global people

    try:
        req = request.Request(url, headers=hds[np.random.randint(0,len(hds))])
        source_code = request.urlopen(req).read()
        plain_text=str(source_code).replace('<br>','').replace('<br/>','').replace('<br />','').replace('\\n','')
    except (error.HTTPError, error.URLError) as e:
        print_builder_lists_excel(builder_list,0)
        print ('Access denied.')
        print (e)

    soup = BeautifulSoup(plain_text)

    try:
        ul = soup.find('ul', {'class':'member-meta'}).find_all('li')
    except:
        print('Company data not found: '+name)
        return detail

    for li in ul:
        if(li.find('strong') == None):
            try:
                [s.extract() for s in li('strong')]
                content = li.string.strip()
            except:
                print('Wrong type address')
                continue
            else:
                detail[5]=content
        elif(li.find('strong').string == 'Builders Reg No: '):
            try:
                [s.extract() for s in li('strong')]
                content = li.string.strip()
            except:
                print('Wrong type Reg No.')
                continue
            else:
                detail[0]=content
        elif(li.find('strong').string == 'Areas:'):
            try:
                [s.extract() for s in li('strong')]
                content = li.string.strip()
            except:
                print('Wrong type area')
                continue
            else:
                detail[1]=content
        elif(li.find('strong').string == 'Contact:'):
            try:
                [s.extract() for s in li('strong')]
                content = li.string.strip()
            except:
                print('Wrong type contact')
                continue
            else:
                if content != '':
                    people.append([content,name])
                    detail[2]=content
        elif(li.find('strong').string == 'T:'):
            try:
                [s.extract() for s in li('strong')]
                content = li.string.strip()
            except:
                print('Wrong type T')
                continue
            else:
                detail[3]=content
        elif(li.find('strong').string == 'F:'):
            try:
                [s.extract() for s in li('strong')]
                content = li.string.strip()
            except:
                print('Wrong type F')
                continue
            else:
                detail[4]=content
    try:
        web = soup.find('a', {'class':'btn btn-primary btn-uppercase m-b-1'}).get('href')
        detail[6]=web
    except:
        print('No web')

    divl = soup.find_all('div', {'class':'m-b-3'})
    for item in divl:
        if(item.find('h3') != None and item.find('h3').string == 'Key Projects'):
            item_list=[]
            li=item.find_all(['li','p'])
            for i in li:
                try:
                    if i.string.strip() != None:
                        item_list.append(i.string.strip())
                except:
                    continue
            detail[7]=','.join(item_list)
        elif(item.find('h3') != None and item.find('h3').string == 'Awards'):
            item_list=[]
            li=item.find_all(['li','p'])
            for i in li:
                try:
                    if i.string.strip() != None:
                        item_list.append(i.string.strip())
                except:
                    continue
            detail[8]=','.join(item_list)

    stuff = soup.find('h4', {'class':'primary'})
    if stuff != None and len(stuff.find_next_siblings('ul')) != 0:
        for person in stuff.find_next_siblings('ul')[0].find_all('li'):
            try:
                if person.string.strip() != '':
                    people.append([len(people),person.string.strip(),name])
            except:
                continue
    elif stuff != None and stuff.find_next_siblings('p') != None:
        for person in stuff.find_next_siblings('p'):
            try:
                if person.string.strip() != '':
                    people.append([len(people),person.string.strip(),name])
            except:
                continue
        
    return detail

def print_builder_lists_excel(builder_list,page_num):
    wb=Workbook(write_only=True)
    global people
    people = list(set([tuple(t) for t in people]))

    ws=wb.create_sheet('builder')
    ws2=wb.create_sheet('people')
    for bl in builder_list:
       ws.append([bl[0],bl[1],bl[2],bl[3],bl[4],bl[5],bl[6],bl[7],bl[8],bl[9],bl[10]])
    for person in people:
       ws2.append([person[0],person[1],person[2]])
    save_path='builder list '+time.strftime("%Y-%m-%d", time.localtime())+' page '+str(page_num)+'.xlsx'
    wb.save(save_path)

if __name__=='__main__':
    builder_lists=builder_spider()